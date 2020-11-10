import os
import sys

import openpyxl as xl

from take_data import parse_raw_data


def enter_data_to_excel(parsed_data, export_folder='excel'):
    for file_name, file_data in parsed_data.items():
        out_file = file_name.split('\\')[-1][:-4]

        print(out_file, file_data)

        wb = xl.Workbook()
        ws = wb.active
        wb.title = out_file

        cap = False  # Наличие шапки(в первую строку)
        if not cap:
            for block in range(1, 9):
                ws.cell(1, block, (['№'] + file_data['header'])[block - 1])

        for number, d in enumerate(file_data['data']):
            ws.cell(int(number) + 2, 1, number + 1)

            for x_instr in range(1, 8):
                ws.cell(int(number) + 2, x_instr + 1, d[x_instr - 1])

        ws.cell(1, 9, "max number")
        ws.cell(1, 10, number + 1)
        wb.save(f'{os.path.join(export_folder, out_file)}.xlsx')
        wb.close()


def main(args):
    if len(args) == 1:
        print('Error: data source path expected as an argument, abort.')
        return

    path = args[1]

    res = parse_raw_data(path)

    enter_data_to_excel(res)


if __name__ == '__main__':
    main(sys.argv)
