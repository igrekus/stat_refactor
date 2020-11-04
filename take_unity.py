import take_data as td
import openpyxl as xl
import os


def take_data_excel():
    if 'excel' not in os.getcwd():
        os.chdir('excel')
        Datas = os.listdir()

    else:
        Datas = os.listdir()

    if 'unity.xlsx' in Datas:
        wb_unity = xl.load_workbook('unity.xlsx')  # Файл соединенный из других файлов
    else:
        wb_unity = xl.Workbook()

    if len(Datas) == 0:
        return False

    for i in Datas:
        if i == 'unity.xlsx':
            continue
        wb_buff = xl.load_workbook(i)
        ws_buff = wb_buff.worksheets[0]
        ws_unity = wb_unity.create_sheet(i)
        for row in ws_buff:
            for cell in row:
                ws_unity[cell.coordinate].value = cell.value
                
        wb_unity.save('unity.xlsx')


take_data_excel()
