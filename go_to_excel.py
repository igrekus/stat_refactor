import take_data as td
import openpyxl as xl
import os

print(os.listdir())
data_txt = td.parse_raw_data()
# data_excel = td.take_data_exel()
cap_name = '№	f, MHz	P, dBm	IG, mA	ID, A	Gain, dB	КПД, %	Pвых, W'.split('\t')


def enter_data_to_excel(data):
    os.chdir(str(chdir_up(os.getcwd())) + '\\excel')
    print(os.getcwd())
    # chdir_up(os.getcwd())
    for i in data.keys():

        wb = xl.Workbook()

        print(i[:-4], data[i])
        meas = data[i]
        ws = wb.active
        wb.title = i[:-4]
        cap = False  # Наличие шапки(в первую строку)
        for block in range(1, 9):
            if not cap:
                ws.cell(1, block, cap_name[block - 1])
        for number in meas.keys():
            ws.cell(int(number) + 1, 1, number)
            for x_instr in range(1, 8):
                # print('test')
                ws.cell(int(number) + 1, x_instr + 1, meas[number][x_instr - 1])
        ws.cell(1, 9, "max number")
        ws.cell(1, 10, number)
        wb.save(i[:-4] + '.xlsx')
        wb.close()


def chdir_up(path):
    path = path.split('\\')
    # print(path)
    path = '\\'.join(path[:-1])
    # print(path)
    return path


# def take_data_excel():
#     if 'excel' not in os.getcwd():
#         os.chdir('excel')
#         Datas = os.listdir()
#
#     else:
#         Datas = os.listdir()
#     wb_unity = xl.Workbook()   # Файл соединенный из других файлов
#
#     if len(Datas) == 0:
#         return False
#
#     for i in Datas:
#         if i == 'unity.xlsx':
#             continue
#         wb_buff = xl.load_workbook(i)
#         print(i['Sheet'])
#         ws_buff = i[wb_buff.sheetnames[0]]
#
#         # # ws_buff = wb_buff.active()
#         # ws_unity = wb_unity.copy_worksheet(wb_buff.active())
#         wb_unity.save('unity.xlsx')





enter_data_to_excel(data_txt)
# take_data_excel()
