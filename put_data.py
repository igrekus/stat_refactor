import itertools
import os
import openpyxl

from openpyxl.chart import LineChart, Reference


def export_to_excel(parsed_data, export_folder='excel'):
    wb_compiled = openpyxl.Workbook()
    ws_plots = wb_compiled.active

    for file_name, file_data in parsed_data.items():
        out_file = file_name.split('\\')[-1][:-4]

        wb = openpyxl.Workbook()
        ws = wb.active
        wb.title = out_file

        ws_compiled = wb_compiled.create_sheet(out_file)

        header = ['№'] + file_data['header'] + ['max number'] + [len(file_data['data'])]
        ws.append(header)
        ws_compiled.append(header)

        for number, d in enumerate(file_data['data']):
            row = [number] + d
            ws.append(row)
            ws_compiled.append(row)

        path = os.path.join(export_folder, out_file)
        print(f'writing excel {path}')
        wb.save(f'{path}.xlsx')
        wb.close()

    print('collecting stats')
    first = next(iter(parsed_data.values()))
    freqs = [f for f, *rest in first['data']]

    header = ['f, GHz', *[f'sheet {i + 1}' for _ in range(len(first['data'][0]) - 1) for i in range(len(parsed_data))]]
    footer = ['', *[h for h in first['header'][1:] for _ in range(len(parsed_data))]]

    data = list(zip(*[v['data'] for v in parsed_data.values()]))
    data1 = [[vs[1:] for vs in d] for d in data]
    data2 = [list(itertools.chain.from_iterable(list(zip(*d)))) for d in data1]
    data3 = [[f] + v for f, v in zip(freqs, data2)]

    ws_plots.append(header)
    for row in data3:
        ws_plots.append(row)
    ws_plots.append(footer)

    base_row = 2
    base_col = 2
    sets = len(parsed_data)
    rows = len(freqs)
    data_blocks = [
        [[base_row, base_col + i * sets], [base_row + rows - 1, base_col + sets + i * sets - 1]]
        for i in range(len(first['header']) - 1)
    ]

    for block in data_blocks:

        (row1, col1), (row2, col2) = block

        chart = LineChart()
        chart.title = first['header'][1:][0]
        chart.style = 2

        dt = Reference(ws_plots, min_row=row1 - 1, min_col=col1, max_row=row2, max_col=col2)
        xs = Reference(ws_plots, min_row=2, min_col=1, max_row=row2)

        chart.add_data(dt, titles_from_data=True)
        chart.set_categories(xs)

        anchor = str(Reference(ws_plots, min_row=10, min_col=col1 - 1)).replace("'Sheet'!$", '').replace('$', '')
        ws_plots.add_chart(chart, anchor=anchor)

    print('writing collected sheet')
    wb_compiled.save(f'{os.path.join(export_folder, "compiled")}.xlsx')
    wb_compiled.close()
